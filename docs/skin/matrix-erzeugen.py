"""Erzeugt Skin-Matrix.xlsx aus Skin.java.

Liest jede CSS-Regel, die Skin.buildCss erzeugt, und schreibt sie als eine Zeile je
(Selektor · Property · Feld) in die Tabelle. Selbst ergaenzte Spalten ab Spalte I bleiben erhalten.

    python docs/skin/matrix-erzeugen.py

Braucht openpyxl:  python -m pip install openpyxl
"""
import re
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

WURZEL = Path(__file__).resolve().parents[2]
SKIN = WURZEL / "src/main/java/app/shared/skin/Skin.java"
PROPS = WURZEL / "src/main/java/app/shared/skin/SkinProperties.java"
ZIEL = Path(__file__).resolve().parent / "Skin-Matrix.xlsx"

KOPF = ["Komponente", "Bereich", "Zustand", "Selektor", "Property", "Feld", "Vorgabe", "Herleitung"]

# --------------------------------------------------------------------------- Vorgaben
# Woher ein Feld seinen Wert bekommt, wenn die properties-Datei schweigt. Handgepflegt:
# die Ableitungen stehen im Vorgaben-Durchlauf am Anfang von Skin.buildCss.
VORGABE = {
    "textColor": "Pflicht", "incorrectTextColor": "Pflicht", "correctColor": "Pflicht",
    "incorrectColor": "Pflicht", "markedColor": "Pflicht", "activeComponentBgColor": "Pflicht",
    "activeComponentHoverColor": "Pflicht", "disabledComponentBgColor": "Pflicht",
    "displayTextBgColor": "Pflicht", "menuBarBackground": "Pflicht", "borderColor": "Pflicht",
    "thinBorderColor": "Pflicht", "borderSmallComponent": "Pflicht",
    "borderMediumComponent": "Pflicht", "borderBigComponent": "Pflicht",
    "font": "Pflicht", "smallFont": "Pflicht",
    "textActiveComponentColor": "← textColor",
    "playFieldBackground": "← menuBarBackground",
    "menuBarHoverBackground": "← contrastingShade(menuBarBackground, 20)",
    "borderShapeColor": "← borderColor",
    "displayTextQuestionBgColor": "← displayTextBgColor",
    "displayTextProgressBgColor": "← displayTextBgColor",
    "displayTextHistoryBgColor": "← displayTextBgColor",
    "displayTextClockBgColor": "← displayTextBgColor",
    "clockPausedTextColor": "← textColor halb zu displayTextClockBgColor",
    "answerSlotWaitingBgColor": "← activeComponentBgColor 10% zu playFieldBackground",
    "clockFont": "← font-Familie, dreifache Größe",
    "menuButtonPadding": "← font × 0,3 / 0,4",
    "menuItemPadding": "← font × 0,1 / 0,5",
    "dashBoardTileTopFontSize": "← font × 4",
    "dashBoardTileBottomFontSize": "← font × 1",
    "mcCorrectTextColor": "entfällt", "mcIncorrectTextColor": "entfällt",
    "shapeMapColor0": "entfällt", "shapeMapColor1": "entfällt",
    "activeBorderColor": "entfällt", "componentShadow": "entfällt",
    "toEliminateColor": "← disabledComponentBgColor", "mcResultBorderWidth": "0",
    "thinBorderWidth": "1", "activeBorderWidth": "2", "imageMapShapeBorderWidth": "2",
    "imageMapLineShapeInnerWidth": "12", "imageMapShapeMarkedOuterWidth": "7",
    "imageMapShapeMarkedInnerWidth": "4",
    "shapeMapStandardBorderWidth": "1.8 — aus der Datei nicht setzbar",
    "mcResultTintPercent": "12",
    "shapeMapFederalStateBorderWidth": "2.8 — aus der Datei nicht setzbar",
    "dashBoardTileTopHeight": "250",
    "dashBoardTileBottomHeight": "100",
    "moviePosterWidth": "154", "imageLabelBgColor": "vollständig durchsichtig",
    "chartRootPadding": '"50px 50px 50px 50px"',
}

# --------------------------------------------------------------------------- Selektor -> Komponente
# Die erste passende Regel gewinnt. Faellt ein Selektor durch, meldet das Skript es.
MAP = [
    (r"^\.root$", "Alles (Wurzel)", "überall"),
    (r"^\.text$", "Jeder Text", "überall"),
    (r"^\.my-map-shape", "Shape-Karte — Form", "Lernen"),
    (r"^\.my-shape-map-pane", "Shape-Karte — Fläche", "Lernen"),
    (r"^\.layer-(neighbor|water)$", "Shape-Karte — Deko", "Lernen"),
    (r"^\.my-image-map-shape", "Bild-Karte — Form", "Lernen"),
    (r"^\.my-image-map-pane", "Bild-Karte — Fläche", "Lernen"),
    (r"^\.(river )?\.?(first|second)$", "Bild-Karte — Klickschicht", "Lernen"),
    (r"^\.river ", "Bild-Karte — Klickschicht", "Lernen"),
    (r"^\.my-image-(background|border)-layer$", "Bild-Panel", "Lernen"),
    (r"^\.my-answer-slot", "Antwortfeld (Fast Write)", "Lernen"),
    (r"^\.my-mc-button", "Antwortknopf (Multiple Choice)", "Lernen"),
    (r"^\.my-info-label\.clock", "Uhr", "Lernen"),
    (r"^\.my-info-label", "Info-Feld (Frage/Fortschritt/Verlauf/Uhr)", "Lernen"),
    (r"^\.my-icon-button", "Icon-Knopf (Zurück)", "Lernen"),
    (r"^\.dashboard-tile", "Dashboard-Kachel", "Dashboard"),
    (r"^\.suite-card-list", "Kartenliste", "Tagebuch + Film"),
    (r"^\.diary-card", "Tagebuchkarte", "Tagebuch"),
    (r"^\.diary-viewer", "Tagebuch — Layout", "Tagebuch"),
    (r"^\.tag-chip-remove", "Schlagwort — Löschknopf", "Tagebuch"),
    (r"^\.suggestion-box", "Vorschlagsliste", "Tagebuch"),
    (r"^\.movie-comment-popup", "Film — Kommentarfenster", "Film"),
    (r"^\.movie-card", "Filmkarte", "Film"),
    (r"^\.movie-viewer", "Film — Layout", "Film"),
    (r"^\.chart-bar", "Diagramm — Balken", "Statistik"),
    (r"^\.chart", "Diagramm", "Statistik"),
    (r"^\.my-table-view", "Tabelle", "Statistik"),
    (r"^\.dialog-pane|^\.my-dialog", "Dialog", "Dialog"),
    (r"^\.date-picker", "Datumsauswahl", "Dialog"),
    (r"^\.spinner", "Zahlenfeld", "Dialog"),
    (r"^\.combo-box", "Auswahlliste", "Dialog"),
    (r"^\.(box|check-box)", "Checkbox", "Dialog"),
    (r"^\.text-area", "Textbereich", "Dialog"),
    (r"^\.text-field", "Eingabefeld", "überall"),
    (r"^\.button", "Knopf", "überall"),
    (r"^\.menu-bar|^\.menu-button|^\.menu-item|^\.menu:|^\.context-menu|^\.my-spacer",
     "Menü", "überall"),
    (r"^\.my-header", "Titelleiste des Hauptfensters", "überall"),
    (r"^\.scroll-bar", "Bildlaufleiste", "überall"),
]


# --------------------------------------------------------------------------- Sonderzeilen
# Erfundene Zeilen. Sie stehen fuer Faelle, in denen das Fehlen einer Regel die Aussage ist —
# die kann der Parser nicht finden, gesucht werden sie trotzdem. Als "(kein Zustand)" erkennbar.
SONDERZEILEN = [
    ["Shape-Karte — Form", "Lernen", "(kein Zustand)", "—", "-fx-fill", "—",
     "keine Füllung",
     "Formen ohne Pseudoklasse: unsichtbar und im Inneren unklickbar, siehe ShapeMapPane"],
]


def ohne_kommentare(text):
    """Block- und Zeilenkommentare raus, String-Literale bleiben unangetastet."""
    zeilen, im_block = [], False
    for s in text.split("\n"):
        if im_block:
            if "*/" in s:
                s, im_block = s.split("*/", 1)[1], False
            else:
                zeilen.append("")
                continue
        if "/*" in s:
            kopf, rest = s.split("/*", 1)
            if "*/" in rest:
                s = kopf + rest.split("*/", 1)[1]
            else:
                s, im_block = kopf, True
        raus, quote, i = [], False, 0
        while i < len(s):
            c = s[i]
            if c == '"' and (i == 0 or s[i - 1] != "\\"):
                quote = not quote
            if not quote and c == "/" and i + 1 < len(s) and s[i + 1] == "/":
                break
            raus.append(c)
            i += 1
        zeilen.append("".join(raus))
    return zeilen


def literal(ausdruck):
    """Konkatenation reiner String-Literale zu einem Text zusammenziehen."""
    if re.fullmatch(r'\s*"[^"]*"(\s*\+\s*"[^"]*")*\s*', ausdruck):
        return "".join(re.findall(r'"([^"]*)"', ausdruck))
    return ausdruck.strip()


def regeln_lesen(zeilen):
    """Jede erzeugte CSS-Regel als (Methode, Selektor, Property, Java-Ausdruck)."""
    methode_re = re.compile(r"private\s+void\s+(add\w+Styles)\s*\(")
    start_re = re.compile(r"\.start\(\s*(.+?)\s*\)\s*$")
    add_re = re.compile(r'\.add(?:IfSet)?\(\s*"(-fx-[\w-]+)"\s*,\s*(.+?)\s*\)\s*$')
    effect_re = re.compile(r"\.effect\(\s*(.+?)\s*\)\s*$")
    aktiv_re = re.compile(r"\.ring\(\s*(.+?)\s*\)\s*$")
    inaktiv_re = re.compile(r"\.resetRing\(\s*(.+?)\s*\)\s*$")
    rule_re = re.compile(r'\.rule\(\s*(.+?)\s*,\s*"(-fx-[\w-]+)"\s*,\s*(.+?)\s*\)\s*;')

    treffer, methode, selektor, puffer = [], "?", None, ""
    for s in zeilen:
        m = methode_re.search(s)
        if m:
            methode = m.group(1)
            continue
        if not s.strip():
            continue
        puffer = (puffer + " " + s.strip()).strip() if puffer else s.strip()
        if puffer.count("(") - puffer.count(")") > 0:      # mehrzeiliger Aufruf
            continue
        zeile, puffer = puffer, ""

        m = rule_re.search(zeile)
        if m:
            treffer.append((methode, literal(m.group(1)), m.group(2), m.group(3).strip()))
            continue
        m = start_re.search(zeile)
        if m:
            selektor = literal(m.group(1))
            continue
        if ".end()" in zeile:
            selektor = None
        if not selektor:
            continue
        m = add_re.search(zeile)
        if m:
            treffer.append((methode, selektor, m.group(1), m.group(2).strip()))
            continue
        m = effect_re.search(zeile)
        if m:
            treffer.append((methode, selektor, "-fx-effect", m.group(1).strip()))
            continue
        m = aktiv_re.search(zeile)
        if m:
            a = [x.strip() for x in m.group(1).split(",")]
            for prop, wert in (("-fx-border-color", a[0]), ("-fx-border-width", a[1]),
                               ("-fx-background-insets", a[1])):
                treffer.append((methode, selektor, prop, wert))
            continue
        m = inaktiv_re.search(zeile)
        if m:
            a = [x.strip() for x in m.group(1).split(",")]
            for prop, ruf in (("-fx-border-color", ".color()"), ("-fx-border-width", ".width()"),
                              ("-fx-background-insets", ".width()")):
                treffer.append((methode, selektor, prop, a[1] + ruf))
    return treffer


def main():
    quelle = SKIN.read_text(encoding="utf-8")
    felder = set(re.findall(r"protected\s+(?:static\s+)?[\w.<>\[\]]+\s+(\w+)\s*[=;]",
                            PROPS.read_text(encoding="utf-8")))

    # Zwischenvariablen je Methode aufloesen — 'padding' heisst in zwei Methoden Verschiedenes.
    alias, teile = {}, re.split(r"private\s+void\s+(add\w+Styles)\s*\(", quelle)
    for k in range(1, len(teile), 2):
        lokal = {}
        for m in re.finditer(
                r"^[ \t]*(?:final\s+)?(?:BorderParams|Insets|Color|double|String|int)\s+(\w+)\s*=\s*(.+?);",
                teile[k + 1], re.M | re.S):
            lokal.setdefault(m.group(1), m.group(2).strip())
        alias[teile[k]] = lokal

    def aufloesen(ausdruck, methode):
        lokal = alias.get(methode, {})
        for _ in range(5):
            neu = ausdruck
            for name, wert in lokal.items():
                neu = re.sub(r"\b" + re.escape(name) + r"\b", "(" + wert + ")", neu)
            # Methode statt Variable — steht in SkinProperties und rechnet aus der Schrift.
            # Muss in der Schleife stehen: erst loest sich die Variable zum Aufruf auf, dann der Aufruf.
            neu = neu.replace("mcLineSpacingSqueezed()", "(font.getSize() * -0.4)")
            if neu == ausdruck:
                break
            ausdruck = neu
        return ausdruck

    def felder_in(ausdruck):
        gefunden = sorted((f for f in felder if re.search(r"\b" + re.escape(f) + r"\b", ausdruck)),
                          key=lambda f: -len(f))
        ergebnis = []
        for f in gefunden:                       # nur den laengsten Treffer je Stelle behalten
            if not any(f != g and f in g for g in ergebnis):
                ergebnis.append(f)
        return ergebnis

    def komponente(sel):
        for muster, name, bereich in MAP:
            if re.search(muster, sel):
                return name, bereich
        return None, None

    def zustand(sel):
        z = [x for x in re.findall(r":([a-z-]+)", sel) if x not in ("left", "bottom", "right")]
        return ", ".join(":" + x for x in z)

    zeilen, offen = [], set()
    for methode, sel, prop, ausdruck in regeln_lesen(ohne_kommentare(quelle)):
        paare = [(sel, ausdruck)]
        if "type.styleClass()" in sel:           # die Schleife ueber TextLabelType
            paare = [(".my-info-label." + t.lower(), "displayText" + t + "BgColor")
                     for t in ("Question", "Progress", "History", "Clock")]
        for s, a in paare:
            komp, bereich = komponente(s)
            if komp is None:
                offen.add(s)
                komp, bereich = "?", "?"
            herleitung = re.sub(r"\s+", " ", a.strip())
            for f in felder_in(aufloesen(a, methode)) or [None]:
                zeilen.append([komp, bereich, zustand(s), s, prop, f or "—",
                               VORGABE.get(f, "fest im Code" if f is None else "?"),
                               "" if f and herleitung == f else herleitung[:90]])

    for s in sorted(offen):
        print("Selektor ohne Komponente:", s)
    for f in sorted({r[5] for r in zeilen if r[6] == "?"}):
        print("Feld ohne Vorgabe-Eintrag:", f)

    zeilen.extend(SONDERZEILEN)
    zeilen.sort(key=lambda r: (r[1], r[0], r[3], r[4]))

    # Schreiben. Spalten ab I gehoeren dem Menschen und bleiben stehen.
    if ZIEL.exists():
        wb = load_workbook(ZIEL)
        ws = wb["Matrix"] if "Matrix" in wb.sheetnames else wb.active
        eigene = {c: [ws.cell(r, c).value for r in range(1, ws.max_row + 1)]
                  for c in range(len(KOPF) + 1, ws.max_column + 1)}
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Matrix"
        eigene = {}

    ws.append(KOPF)
    for r in zeilen:
        ws.append(r)
    for c, werte in eigene.items():
        for i, v in enumerate(werte, start=1):
            ws.cell(i, c).value = v

    fuellung = PatternFill("solid", fgColor="1C2430")
    for c in range(1, ws.max_column + 1):
        z = ws.cell(1, c)
        z.font = Font(bold=True, color="FFFFFF")
        z.fill = fuellung
        z.alignment = Alignment(vertical="center")
    for c, breite in zip(range(1, 9), (36, 12, 22, 46, 24, 30, 34, 46)):
        ws.column_dimensions[get_column_letter(c)].width = breite
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ws.max_column), ws.max_row)
    wb.save(ZIEL)

    print("%s: %d Zeilen, %d Selektoren, %d Felder"
          % (ZIEL.name, len(zeilen), len({r[3] for r in zeilen}),
             len({r[5] for r in zeilen if r[5] != "—"})))


if __name__ == "__main__":
    main()
