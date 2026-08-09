"""Prueft Skin-Matrix.xlsx gegen den Quellcode.

    python docs/skin/matrix-pruefen.py

Drei Fragen:
  1. Ist jeder Selektor erfasst, den Skin.java erzeugt?
  2. Gibt es jeden Feldnamen der Tabelle wirklich in SkinProperties?
  3. Welche Felder stehen nicht in der Tabelle? Die gehoeren nach Skin-Felder.md.

Braucht einen Build (target/classes) fuer Frage 2 — ohne den wird sie uebersprungen.
"""
import os
import re
import subprocess
from pathlib import Path
from openpyxl import load_workbook

import importlib.util
import sys

sys.dont_write_bytecode = True   # kein __pycache__ neben der Doku

HIER = Path(__file__).resolve().parent
WURZEL = HIER.parents[1]
SKIN = WURZEL / "src/main/java/app/shared/skin/Skin.java"
PROPS = WURZEL / "src/main/java/app/shared/skin/SkinProperties.java"
KLASSE = WURZEL / "target/classes/app/shared/skin/SkinProperties.class"
XLSX = HIER / "Skin-Matrix.xlsx"

spec = importlib.util.spec_from_file_location("erzeugen", HIER / "matrix-erzeugen.py")
erzeugen = importlib.util.module_from_spec(spec)
spec.loader.exec_module(erzeugen)

ws = load_workbook(XLSX)["Matrix"]
daten = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
sonder = [r for r in daten if r[2] == "(kein Zustand)"]      # erfundene Zeilen, siehe Erzeuger
echt = [r for r in daten if r not in sonder]
sel_tabelle = {r[3] for r in echt}
felder_tabelle = {r[5] for r in echt if r[5] != "—"}

# --- 1. Vollzaehligkeit der Selektoren
zeilen = erzeugen.ohne_kommentare(SKIN.read_text(encoding="utf-8"))
sel_code = {s for _, s, _, _ in erzeugen.regeln_lesen(zeilen)}
dynamisch = {s for s in sel_code if "styleClass()" in s}
fehlend = {s for s in sel_code - dynamisch if s not in sel_tabelle}
print("1) Selektoren im Code:", len(sel_code), "| in der Tabelle:", len(sel_tabelle))
print("   Fehlend:", fehlend or "keine",
      "| dynamisch erzeugt (aufgefaechert):", len(dynamisch))

# --- 2. Feldnamen
# Quelle UND Bytecode: der Build kann aelter sein als die Quelle, dann fehlen dort neue Felder.
# Namen mit grossem Anfangsbuchstaben sind Typen (verschachtelte Enums), keine Felder.
felder_klasse = {f for f in re.findall(
        r"protected\s+(?:static\s+)?[\w.<>\[\]]+\s+(\w+)\s*[=;]", PROPS.read_text(encoding="utf-8"))
        if f[:1].islower()}
if KLASSE.exists() and os.environ.get("JAVA_HOME"):
    javap = Path(os.environ["JAVA_HOME"]) / "bin/javap.exe"
    ausgabe = subprocess.run([str(javap), "-p", str(KLASSE)], capture_output=True, text=True).stdout
    felder_klasse |= {m.group(1) for m in
                      (re.search(r"(\w+);\s*$", z) for z in ausgabe.split("\n"))
                      if m and m.group(1)[:1].islower()}
print("2) Feldnamen der Tabelle, die es nicht gibt:", felder_tabelle - felder_klasse or "keine")

# --- 3. Was nicht in der Tabelle steht
rest = sorted(felder_klasse - felder_tabelle)
rechtecke = [f for f in rest if re.search(r"Session(\w+Panel|BackButton)$", f)]
bilder = [f for f in rest if re.search(r"(WallpaperName|MapImageName|MapOverlayImageName|"
                                      r"MapInactiveImageName|MapInactiveOverlayImageName|ButtonIcon)$", f)]
sonstige = [f for f in rest if f not in rechtecke and f not in bilder]
print("3) Nicht in der Tabelle:", len(rest),
      "(%d Rechtecke, %d Bild-/Symbolnamen, %d sonstige)" % (len(rechtecke), len(bilder), len(sonstige)))
print("   Sonstige:", ", ".join(sonstige))

print("\nZeilen gesamt:", len(daten),
      "| davon ohne Feld (fest im Code):", sum(1 for r in echt if r[5] == "—"),
      "| erfundene Sonderzeilen:", len(sonder))
